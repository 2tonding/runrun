import os
import io
import json
import re
import secrets
import httpx
import redis
from fastapi import FastAPI, Request, Depends, HTTPException, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from anthropic import Anthropic
from datetime import datetime, timedelta

# ============================================================
# CONFIGURACAO
# ============================================================
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
ZAPI_INSTANCE_ID  = os.environ.get("ZAPI_INSTANCE_ID")
ZAPI_TOKEN        = os.environ.get("ZAPI_TOKEN")
ZAPI_CLIENT_TOKEN = os.environ.get("ZAPI_CLIENT_TOKEN")
REDIS_URL         = os.environ.get("REDIS_URL")
ADMIN_USER        = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASS        = os.environ.get("ADMIN_PASS", "admin123")
BASE_URL          = os.environ.get("BASE_URL", "https://SEU-DOMINIO.up.railway.app")
AGENT_NAME        = os.environ.get("AGENT_NAME", "PrimeiraMente")
AGENT_MODEL       = os.environ.get("AGENT_MODEL", "claude-haiku-4-5-20251001")
GROQ_API_KEY      = os.environ.get("GROQ_API_KEY")
MP_ACCESS_TOKEN   = os.environ.get("MP_ACCESS_TOKEN")
MP_PUBLIC_KEY     = os.environ.get("MP_PUBLIC_KEY")
MP_PLAN_ID        = os.environ.get("MP_PLAN_ID", "bc34d81de9ba466b8d2693d1a134871c")
LINK_PAGAMENTO    = os.environ.get("LINK_PAGAMENTO", "")  # link externo opcional

AGENT_PROMPT_PADRAO = """Voce e um assistente prestativo e simpatico.
Responda de forma clara, direta e em portugues.
No WhatsApp, seja breve — uma ideia por mensagem, no maximo."""

client   = Anthropic(api_key=ANTHROPIC_API_KEY)
app      = FastAPI()
security = HTTPBasic()

# ============================================================
# REDIS
# ============================================================
r = redis.from_url(REDIS_URL, decode_responses=True)

# ============================================================
# AUTENTICACAO ADMIN
# ============================================================
def verificar_admin(credentials: HTTPBasicCredentials = Depends(security)):
    usuario_ok = secrets.compare_digest(credentials.username.encode(), ADMIN_USER.encode())
    senha_ok   = secrets.compare_digest(credentials.password.encode(), ADMIN_PASS.encode())
    if not (usuario_ok and senha_ok):
        raise HTTPException(
            status_code=401,
            detail="Acesso negado",
            headers={"WWW-Authenticate": "Basic"}
        )
    return credentials.username

# ============================================================
# PROMPT — salvo no Redis, editavel pelo painel
# ============================================================
PROMPT_KEY = "config:agent_prompt"

def obter_prompt() -> str:
    prompt = r.get(PROMPT_KEY)
    return prompt if prompt else AGENT_PROMPT_PADRAO

def salvar_prompt(prompt: str):
    r.set(PROMPT_KEY, prompt)

# ============================================================
# ARQUIVOS DE REFERENCIA
# ============================================================
ARQUIVO_PREFIX = "config:arquivo:"

def listar_arquivos() -> list:
    chaves = r.keys(f"{ARQUIVO_PREFIX}*")
    arquivos = []
    for chave in sorted(chaves):
        nome = chave.replace(ARQUIVO_PREFIX, "")
        tamanho = len(r.get(chave) or "")
        arquivos.append({"nome": nome, "tamanho": tamanho})
    return arquivos

def obter_arquivo(nome: str) -> str | None:
    return r.get(f"{ARQUIVO_PREFIX}{nome}")

def salvar_arquivo(nome: str, conteudo: str):
    r.set(f"{ARQUIVO_PREFIX}{nome}", conteudo[:20000])

def apagar_arquivo(nome: str):
    r.delete(f"{ARQUIVO_PREFIX}{nome}")

def injetar_arquivos_no_prompt(prompt: str) -> str:
    referencias = re.findall(r'\[([a-zA-Z0-9_\-]+)\]', prompt)
    for nome in referencias:
        conteudo = obter_arquivo(nome)
        if conteudo:
            prompt = prompt.replace(f"[{nome}]", f"\n\n=== CONTEUDO DE '{nome}' ===\n{conteudo}\n=== FIM DE '{nome}' ===\n")
    return prompt

# ============================================================
# ASSINATURAS
# ============================================================
ASSINATURA_PREFIX = "assinatura:"
CONSULTA_PREFIX   = "consulta:"

def obter_assinatura(telefone: str) -> dict:
    dados = r.get(f"{ASSINATURA_PREFIX}{telefone}")
    if not dados:
        return {"status": "freemium", "plano": "freemium", "telefone": telefone}
    return json.loads(dados)

def salvar_assinatura(telefone: str, dados: dict):
    r.set(f"{ASSINATURA_PREFIX}{telefone}", json.dumps(dados))

def eh_premium(telefone: str) -> bool:
    assinatura = obter_assinatura(telefone)
    if assinatura.get("status") != "ativo":
        return False
    # Verifica vencimento se existir
    expira = assinatura.get("expira")
    if expira:
        try:
            if datetime.fromisoformat(expira) < datetime.now():
                return False
        except Exception:
            pass
    return True

def listar_assinaturas() -> list:
    chaves = r.keys(f"{ASSINATURA_PREFIX}*")
    resultado = []
    for chave in sorted(chaves):
        dados = json.loads(r.get(chave) or "{}")
        resultado.append(dados)
    return resultado

def registrar_interesse_consulta(telefone: str, nome: str):
    dados = {
        "telefone": telefone,
        "nome": nome,
        "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "atendido": False
    }
    r.set(f"{CONSULTA_PREFIX}{telefone}", json.dumps(dados))

def listar_consultas() -> list:
    chaves = r.keys(f"{CONSULTA_PREFIX}*")
    resultado = []
    for chave in sorted(chaves):
        dados = json.loads(r.get(chave) or "{}")
        resultado.append(dados)
    return resultado

def marcar_consulta_atendida(telefone: str):
    dados = json.loads(r.get(f"{CONSULTA_PREFIX}{telefone}") or "{}")
    dados["atendido"] = True
    r.set(f"{CONSULTA_PREFIX}{telefone}", json.dumps(dados))

# ============================================================
# HISTORICO COM REDIS
# ============================================================
HISTORICO_LIMITE = 40

def obter_historico(telefone: str) -> list:
    dados = r.get(f"historico:{telefone}")
    if not dados:
        return []
    return json.loads(dados)[-HISTORICO_LIMITE:]

def salvar_historico(telefone: str, historico: list):
    r.set(f"historico:{telefone}", json.dumps(historico))

def salvar_mensagem(telefone: str, role: str, conteudo: str):
    historico = obter_historico(telefone)
    historico.append({"role": role, "content": conteudo})
    salvar_historico(telefone, historico)

# ============================================================
# PROCESSAMENTO DE MIDIA
# ============================================================

async def transcrever_audio(url_audio: str) -> str:
    if not GROQ_API_KEY:
        return "[Audio recebido, mas GROQ_API_KEY nao configurada]"
    try:
        async with httpx.AsyncClient(timeout=60) as http:
            r_audio = await http.get(url_audio)
            conteudo = r_audio.content
            response = await http.post(
                "https://api.groq.com/openai/v1/audio/transcriptions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                files={"file": ("audio.ogg", conteudo, "audio/ogg")},
                data={"model": "whisper-large-v3", "language": "pt"}
            )
            if response.status_code == 200:
                texto = response.json().get("text", "")
                return f"[Audio transcrito]: {texto}"
            return "[Nao foi possivel transcrever o audio]"
    except Exception as e:
        print(f"ERRO ao transcrever audio: {e}")
        return "[Erro ao processar audio]"


async def extrair_texto_pdf(url_arquivo: str) -> str:
    try:
        import pdfplumber
        async with httpx.AsyncClient(timeout=60) as http:
            r_arquivo = await http.get(url_arquivo)
            conteudo = r_arquivo.content
        with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
            paginas = []
            for i, pagina in enumerate(pdf.pages[:20]):
                texto = pagina.extract_text()
                if texto:
                    paginas.append(f"[Pagina {i+1}]\n{texto}")
        texto_completo = "\n\n".join(paginas)
        return f"[Conteudo do PDF enviado pelo usuario]:\n{texto_completo[:8000]}"
    except Exception as e:
        print(f"ERRO ao ler PDF: {e}")
        return "[Nao foi possivel ler o PDF]"


async def extrair_texto_excel(url_arquivo: str) -> str:
    try:
        import openpyxl
        async with httpx.AsyncClient(timeout=60) as http:
            r_arquivo = await http.get(url_arquivo)
            conteudo = r_arquivo.content
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        linhas_total = []
        for nome_aba in wb.sheetnames[:3]:
            ws = wb[nome_aba]
            linhas_total.append(f"[Aba: {nome_aba}]")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 100:
                    linhas_total.append("... (mais linhas omitidas)")
                    break
                linha = " | ".join(str(c) if c is not None else "" for c in row)
                if linha.strip():
                    linhas_total.append(linha)
        texto_completo = "\n".join(linhas_total)
        return f"[Conteudo da planilha enviada pelo usuario]:\n{texto_completo[:8000]}"
    except Exception as e:
        print(f"ERRO ao ler Excel: {e}")
        return "[Nao foi possivel ler a planilha]"


async def processar_midia(dados: dict) -> str | None:
    audio = dados.get("audio", {})
    if audio and audio.get("audioUrl"):
        return await transcrever_audio(audio["audioUrl"])
    documento = dados.get("document", {})
    if documento:
        url      = documento.get("documentUrl", "")
        filename = documento.get("fileName", "").lower()
        mime     = documento.get("mimeType", "").lower()
        if not url:
            return None
        if "pdf" in mime or filename.endswith(".pdf"):
            return await extrair_texto_pdf(url)
        if any(x in mime for x in ["excel", "spreadsheet", "xlsx", "xls"]) or \
           filename.endswith((".xlsx", ".xls")):
            return await extrair_texto_excel(url)
        return f"[Arquivo recebido: {filename} — tipo nao suportado para leitura automatica]"
    return None

# ============================================================
# CSS DO PAINEL
# ============================================================
CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Arial, sans-serif; background: #f0f2f5; color: #333; }
header { background: #1a1a2e; color: white; padding: 16px 24px; display: flex; align-items: center; gap: 12px; }
header h1 { font-size: 18px; }
header a { color: #aab4ff; text-decoration: none; font-size: 14px; margin-left: auto; }
header a:hover { text-decoration: underline; }
.container { max-width: 860px; margin: 30px auto; padding: 0 20px; }
.card { background: white; border-radius: 12px; padding: 20px; margin-bottom: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
.card h2 { font-size: 16px; margin-bottom: 16px; color: #555; }
a { color: #4f46e5; text-decoration: none; }
a:hover { text-decoration: underline; }
.badge { display: inline-block; border-radius: 20px; padding: 2px 10px; font-size: 12px; margin-left: 8px; }
.badge-premium { background: #fef9c3; color: #b45309; }
.badge-freemium { background: #f1f5f9; color: #64748b; }
.badge-ativo { background: #dcfce7; color: #16a34a; }
.badge-inativo { background: #fee2e2; color: #dc2626; }
.badge-pendente { background: #fef3c7; color: #d97706; }
.aluno-row { display: flex; justify-content: space-between; align-items: center; padding: 12px 0; border-bottom: 1px solid #f0f0f0; }
.aluno-row:last-child { border-bottom: none; }
.aluno-info { font-size: 12px; color: #999; margin-top: 4px; max-width: 540px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.btn { display: inline-block; padding: 6px 14px; border-radius: 8px; font-size: 13px; cursor: pointer; border: none; text-decoration: none; }
.btn-primary { background: #4f46e5; color: white; }
.btn-primary:hover { background: #4338ca; text-decoration: none; color: white; }
.btn-success { background: #dcfce7; color: #16a34a; }
.btn-success:hover { background: #bbf7d0; text-decoration: none; }
.btn-danger { background: #fee2e2; color: #dc2626; }
.btn-danger:hover { background: #fecaca; text-decoration: none; }
.btn-warning { background: #fef3c7; color: #d97706; }
.btn-warning:hover { background: #fde68a; text-decoration: none; }
.back { display: inline-block; margin-bottom: 16px; font-size: 14px; }
.total { font-size: 13px; color: #888; margin-bottom: 16px; }
.chat { display: flex; flex-direction: column; gap: 10px; }
.msg { display: flex; flex-direction: column; max-width: 80%; }
.msg.usuario { align-self: flex-end; align-items: flex-end; }
.msg.agente { align-self: flex-start; align-items: flex-start; }
.label { font-size: 11px; color: #aaa; margin-bottom: 3px; padding: 0 6px; }
.balao { padding: 10px 14px; border-radius: 16px; font-size: 14px; line-height: 1.6; }
.usuario .balao { background: #dcf8c6; border-bottom-right-radius: 4px; }
.agente .balao { background: #f8f8f8; border-bottom-left-radius: 4px; box-shadow: 0 1px 2px rgba(0,0,0,0.08); }
textarea { width: 100%; padding: 12px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; font-family: monospace; line-height: 1.6; resize: vertical; min-height: 300px; }
textarea:focus { outline: none; border-color: #4f46e5; box-shadow: 0 0 0 2px rgba(79,70,229,0.1); }
input[type=text], input[type=number] { width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 8px; font-size: 14px; }
.success { background: #dcfce7; color: #16a34a; padding: 10px 16px; border-radius: 8px; margin-bottom: 16px; font-size: 14px; }
.nav { display: flex; gap: 12px; margin-bottom: 20px; flex-wrap: wrap; }
.nav a { padding: 8px 16px; border-radius: 8px; background: white; font-size: 14px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
.nav a.ativo { background: #4f46e5; color: white; }
.nav a:hover { text-decoration: none; background: #e0e7ff; }
.nav a.ativo:hover { background: #4338ca; }
.stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-bottom: 20px; }
.stat { background: white; border-radius: 12px; padding: 16px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
.stat .num { font-size: 28px; font-weight: bold; color: #4f46e5; }
.stat .label { font-size: 12px; color: #888; margin-top: 4px; }
"""

def base_html(titulo: str, conteudo: str, pagina_ativa: str = "") -> str:
    nav = {
        "usuarios":    ("Usuarios",    "/admin"),
        "assinaturas": ("Assinaturas", "/admin/assinaturas"),
        "consultas":   ("Consultas",   "/admin/consultas"),
        "prompt":      ("Prompt",      "/admin/prompt"),
        "arquivos":    ("Arquivos",    "/admin/arquivos"),
    }
    nav_html = ""
    for chave, (label, url) in nav.items():
        ativo = 'class="ativo"' if pagina_ativa == chave else ""
        nav_html += f'<a href="{url}" {ativo}>{label}</a>'

    return f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{titulo} — {AGENT_NAME}</title>
    <style>{CSS}</style>
</head>
<body>
    <header>
        <span>🧠</span>
        <h1>{AGENT_NAME} — Painel Admin</h1>
        <a href="/admin">Inicio</a>
    </header>
    <div class="container">
        <div class="nav">{nav_html}</div>
        {conteudo}
    </div>
</body>
</html>"""

# ============================================================
# FUNCOES AUXILIARES
# ============================================================

async def enviar_whatsapp(telefone: str, mensagem: str):
    numero_limpo = telefone.replace("+", "").replace("-", "").replace(" ", "")
    if numero_limpo.startswith("55") and len(numero_limpo) == 12:
        numero_limpo = numero_limpo[:4] + "9" + numero_limpo[4:]
    url     = f"https://api.z-api.io/instances/{ZAPI_INSTANCE_ID}/token/{ZAPI_TOKEN}/send-text"
    headers = {"Content-Type": "application/json", "Client-Token": ZAPI_CLIENT_TOKEN}
    payload = {"phone": numero_limpo, "message": mensagem}
    print(f"ENVIANDO para {numero_limpo}")
    async with httpx.AsyncClient(timeout=30) as http:
        response = await http.post(url, headers=headers, json=payload)
        print(f"Z-API STATUS: {response.status_code} | {response.text}")


def obter_link_pagamento(telefone: str) -> str:
    """Gera link de pagamento do Mercado Pago com telefone como referencia."""
    if LINK_PAGAMENTO:
        return LINK_PAGAMENTO
    # Link direto para o plano MP com referencia do usuario
    return f"https://www.mercadopago.com.br/subscriptions/checkout?preapproval_plan_id={MP_PLAN_ID}&back_url={BASE_URL}/pagamento/obrigado&external_reference={telefone}"


async def chamar_claude(telefone: str, mensagem_usuario: str) -> str:
    # Detecta interesse em consulta na mensagem
    palavras_consulta = ["consulta", "agendar", "teleconsulta", "atendimento", "marcar"]
    if any(p in mensagem_usuario.lower() for p in palavras_consulta):
        # Verifica se ja tem nome registrado para consulta
        dados_consulta = r.get(f"{CONSULTA_PREFIX}{telefone}")
        if not dados_consulta:
            # Registra com nome desconhecido por enquanto, sera atualizado
            registrar_interesse_consulta(telefone, "Nome nao informado")
            print(f"INTERESSE CONSULTA registrado: {telefone}")

    salvar_mensagem(telefone, "user", mensagem_usuario)
    historico = obter_historico(telefone)

    hoje       = datetime.now().strftime("%d/%m/%Y")
    dia_semana = ["segunda-feira","terca-feira","quarta-feira","quinta-feira",
                  "sexta-feira","sabado","domingo"][datetime.now().weekday()]

    # Injeta status de assinatura no prompt
    status_usuario = "PREMIUM" if eh_premium(telefone) else "FREEMIUM"
    link_pg = obter_link_pagamento(telefone)

    prompt_base = injetar_arquivos_no_prompt(obter_prompt())
    prompt_base = prompt_base.replace("{STATUS}", status_usuario)
    prompt_base = prompt_base.replace("[LINK_PAGAMENTO]", link_pg)

    system = prompt_base + f"\n\nDATA ATUAL: {dia_semana}, {hoje}\nSTATUS DO USUARIO: {status_usuario}\nLINK DE PAGAMENTO: {link_pg}"

    resposta = client.messages.create(
        model=AGENT_MODEL,
        max_tokens=1024,
        system=system,
        messages=historico
    )

    texto_resposta = resposta.content[0].text
    salvar_mensagem(telefone, "assistant", texto_resposta)

    # Detecta se o bot mencionou interesse em consulta na resposta
    if "registrar seu interesse" in texto_resposta.lower() or "vou registrar" in texto_resposta.lower():
        # Tenta extrair nome da ultima mensagem do usuario
        registrar_interesse_consulta(telefone, mensagem_usuario[:50])

    return texto_resposta

# ============================================================
# ROTAS PUBLICAS
# ============================================================

@app.get("/")
def status():
    return {"status": f"{AGENT_NAME} online"}


@app.get("/pagamento", response_class=HTMLResponse)
async def pagina_pagamento(ref: str = ""):
    """Pagina de pagamento com link para o plano MP."""
    telefone = ref or "visitante"
    link = obter_link_pagamento(telefone)
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Assinar PrimeiraMente Premium</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: Arial, sans-serif; background: #f0f9ff; display: flex; align-items: center; justify-content: center; min-height: 100vh; padding: 20px; }}
        .card {{ background: white; border-radius: 16px; padding: 40px; max-width: 460px; width: 100%; box-shadow: 0 4px 24px rgba(0,0,0,0.10); text-align: center; }}
        .emoji {{ font-size: 48px; margin-bottom: 16px; }}
        h1 {{ font-size: 22px; color: #1a1a2e; margin-bottom: 8px; }}
        p {{ font-size: 15px; color: #555; margin-bottom: 24px; line-height: 1.6; }}
        .preco {{ font-size: 36px; font-weight: bold; color: #4f46e5; margin-bottom: 8px; }}
        .preco span {{ font-size: 16px; color: #888; font-weight: normal; }}
        ul {{ text-align: left; margin: 16px 0 28px; padding: 0 8px; list-style: none; }}
        ul li {{ padding: 6px 0; font-size: 14px; color: #444; }}
        ul li::before {{ content: "✓ "; color: #16a34a; font-weight: bold; }}
        .btn {{ display: block; background: #4f46e5; color: white; padding: 16px; border-radius: 10px; font-size: 16px; font-weight: bold; text-decoration: none; }}
        .btn:hover {{ background: #4338ca; }}
        .seguro {{ font-size: 12px; color: #aaa; margin-top: 16px; }}
    </style>
</head>
<body>
    <div class="card">
        <div class="emoji">🧠</div>
        <h1>PrimeiraMente Premium</h1>
        <p>Orientacao especializada em psiquiatria infantil para pais que querem entender e ajudar seus filhos de verdade.</p>
        <div class="preco">R$ 29,90 <span>/ mes</span></div>
        <ul>
            <li>Orientacao personalizada por comportamento</li>
            <li>Identificacao de sinais de TDAH e TEA</li>
            <li>Estrategias praticas baseadas em ciencia</li>
            <li>Como se preparar para consultas medicas</li>
            <li>Acesso ilimitado 24h pelo WhatsApp</li>
        </ul>
        <a href="{link}" class="btn">Assinar agora — R$ 29,90/mes</a>
        <p class="seguro">🔒 Pagamento seguro via Mercado Pago. Cancele quando quiser.</p>
    </div>
</body>
</html>""")


@app.get("/pagamento/obrigado", response_class=HTMLResponse)
async def pagamento_obrigado(external_reference: str = "", collection_status: str = ""):
    """Pagina de retorno apos pagamento no MP."""
    if collection_status == "approved" and external_reference:
        # Ativa o usuario imediatamente via retorno
        dados = {
            "telefone": external_reference,
            "status": "ativo",
            "plano": "premium",
            "origem": "mercadopago_retorno",
            "data_inicio": datetime.now().isoformat(),
            "expira": (datetime.now() + timedelta(days=35)).isoformat()
        }
        salvar_assinatura(external_reference, dados)
        await enviar_whatsapp(external_reference,
            "Seu acesso Premium esta ativo! 🎉\n\nBem-vindo ao PrimeiraMente Premium. Pode me contar agora o que esta acontecendo com seu filho — estou aqui para ajudar com orientacoes completas 🧠💙")

    return HTMLResponse("""<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <title>Pagamento confirmado</title>
    <style>
        body { font-family: Arial, sans-serif; background: #f0fdf4; display: flex; align-items: center; justify-content: center; min-height: 100vh; }
        .card { background: white; border-radius: 16px; padding: 40px; max-width: 400px; text-align: center; box-shadow: 0 4px 24px rgba(0,0,0,0.08); }
        .emoji { font-size: 56px; margin-bottom: 16px; }
        h1 { color: #16a34a; margin-bottom: 12px; }
        p { color: #555; line-height: 1.6; }
    </style>
</head>
<body>
    <div class="card">
        <div class="emoji">🎉</div>
        <h1>Pagamento confirmado!</h1>
        <p>Seu acesso Premium ja esta ativo.<br>Volte ao WhatsApp — seu assistente esta te esperando!</p>
    </div>
</body>
</html>""")

# ============================================================
# WEBHOOK MERCADO PAGO
# ============================================================

@app.post("/webhook/mercadopago")
async def webhook_mercadopago(request: Request):
    """Recebe notificacoes de pagamento do Mercado Pago."""
    try:
        dados = await request.json()
        print(f"MP WEBHOOK: {json.dumps(dados)[:300]}")

        tipo  = dados.get("type") or dados.get("action", "")
        id_mp = dados.get("data", {}).get("id") or dados.get("id")

        if not id_mp:
            return {"status": "ignorado"}

        # Assinatura criada ou atualizada
        if "subscription" in tipo or "preapproval" in tipo:
            await processar_assinatura_mp(str(id_mp))
        # Pagamento avulso (cobranca recorrente)
        elif "payment" in tipo:
            await processar_pagamento_mp(str(id_mp))

        return {"status": "ok"}
    except Exception as e:
        print(f"ERRO webhook MP: {e}")
        return {"status": "erro", "detalhe": str(e)}


async def processar_assinatura_mp(preapproval_id: str):
    """Busca detalhes da assinatura no MP e ativa/desativa o usuario."""
    if not MP_ACCESS_TOKEN:
        return
    try:
        async with httpx.AsyncClient(timeout=30) as http:
            res = await http.get(
                f"https://api.mercadopago.com/preapproval/{preapproval_id}",
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
            )
            if res.status_code != 200:
                print(f"MP ASSINATURA ERRO: {res.status_code}")
                return
            dados = res.json()

        telefone        = dados.get("external_reference", "")
        status_mp       = dados.get("status", "")  # authorized, paused, cancelled
        proximo_debito  = dados.get("next_payment_date", "")

        if not telefone:
            print("MP: external_reference vazio — nao foi possivel identificar usuario")
            return

        status_local = "ativo" if status_mp == "authorized" else "inativo"
        expira = None
        if proximo_debito:
            try:
                expira = (datetime.fromisoformat(proximo_debito[:10]) + timedelta(days=5)).isoformat()
            except Exception:
                expira = (datetime.now() + timedelta(days=35)).isoformat()

        assinatura = {
            "telefone": telefone,
            "status": status_local,
            "plano": "premium",
            "preapproval_id": preapproval_id,
            "status_mp": status_mp,
            "data_inicio": datetime.now().isoformat(),
            "expira": expira or (datetime.now() + timedelta(days=35)).isoformat()
        }
        salvar_assinatura(telefone, assinatura)
        print(f"ASSINATURA ATUALIZADA: {telefone} → {status_local}")

        if status_local == "ativo":
            await enviar_whatsapp(telefone,
                "Seu acesso Premium esta ativo! 🎉\n\nBem-vindo ao PrimeiraMente Premium. Pode me contar o que esta acontecendo com seu filho — estou aqui para ajudar com orientacoes completas 🧠💙")
        elif status_mp in ("cancelled", "paused"):
            await enviar_whatsapp(telefone,
                "Seu plano Premium foi cancelado. Sentiremos sua falta 💙\n\nSe quiser reativar a qualquer momento, e so me chamar aqui!")

    except Exception as e:
        print(f"ERRO ao processar assinatura MP: {e}")


async def processar_pagamento_mp(payment_id: str):
    """Processa um pagamento recorrente aprovado."""
    if not MP_ACCESS_TOKEN:
        return
    try:
        async with httpx.AsyncClient(timeout=30) as http:
            res = await http.get(
                f"https://api.mercadopago.com/v1/payments/{payment_id}",
                headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
            )
            if res.status_code != 200:
                return
            dados = res.json()

        status_pg  = dados.get("status", "")
        telefone   = dados.get("external_reference", "")

        if not telefone:
            return

        if status_pg == "approved":
            assinatura = obter_assinatura(telefone)
            assinatura["status"] = "ativo"
            assinatura["expira"] = (datetime.now() + timedelta(days=35)).isoformat()
            assinatura["ultimo_pagamento"] = datetime.now().isoformat()
            salvar_assinatura(telefone, assinatura)
            print(f"PAGAMENTO APROVADO: {telefone}")
        elif status_pg in ("rejected", "cancelled"):
            assinatura = obter_assinatura(telefone)
            assinatura["status"] = "inativo"
            salvar_assinatura(telefone, assinatura)
            await enviar_whatsapp(telefone,
                "Tivemos um problema com o pagamento da sua assinatura 😕\n\nPor favor, atualize seu metodo de pagamento para continuar com o acesso Premium.")

    except Exception as e:
        print(f"ERRO ao processar pagamento MP: {e}")

# ============================================================
# PAINEL ADMIN — USUARIOS
# ============================================================

@app.get("/admin", response_class=HTMLResponse)
def painel_admin(admin: str = Depends(verificar_admin)):
    chaves = r.keys("historico:*")

    total_usuarios  = len(chaves)
    total_premium   = len([1 for c in r.keys(f"{ASSINATURA_PREFIX}*")
                           if json.loads(r.get(c) or "{}").get("status") == "ativo"])
    total_consultas = len([1 for c in r.keys(f"{CONSULTA_PREFIX}*")
                           if not json.loads(r.get(c) or "{}").get("atendido")])

    stats = f"""
    <div class="stats">
        <div class="stat"><div class="num">{total_usuarios}</div><div class="label">Usuarios</div></div>
        <div class="stat"><div class="num">{total_premium}</div><div class="label">Premium ativos</div></div>
        <div class="stat"><div class="num">{total_consultas}</div><div class="label">Consultas pendentes</div></div>
    </div>"""

    rows = ""
    for chave in sorted(chaves):
        telefone  = chave.replace("historico:", "")
        historico = obter_historico(telefone)
        total     = len(historico)
        ultima    = historico[-1]["content"][:80] + "..." if historico else "—"
        premium   = eh_premium(telefone)
        badge     = '<span class="badge badge-premium">Premium</span>' if premium else '<span class="badge badge-freemium">Freemium</span>'
        rows += f"""
        <div class="aluno-row">
            <div>
                <div>
                    <a href="/admin/conversa/{telefone}">{telefone}</a>
                    {badge}
                    <span class="badge" style="background:#e0e7ff;color:#4f46e5">{total} msgs</span>
                </div>
                <div class="aluno-info">{ultima}</div>
            </div>
            <a href="/admin/apagar/{telefone}" onclick="return confirm('Apagar historico de {telefone}?')">
                <span class="btn btn-danger">Apagar</span>
            </a>
        </div>"""

    if not rows:
        rows = "<p style='color:#888;padding:20px 0'>Nenhum usuario ainda.</p>"

    conteudo = stats + f"""
    <div class="card">
        <h2>Usuarios ({total_usuarios} total)</h2>
        {rows}
    </div>"""

    return HTMLResponse(base_html("Usuarios", conteudo, "usuarios"))


@app.get("/admin/conversa/{telefone}", response_class=HTMLResponse)
def ver_conversa(telefone: str, admin: str = Depends(verificar_admin)):
    historico = obter_historico(telefone)
    assinatura = obter_assinatura(telefone)
    premium = eh_premium(telefone)

    status_badge = '<span class="badge badge-ativo">Premium ativo</span>' if premium else '<span class="badge badge-freemium">Freemium</span>'
    expira = assinatura.get("expira", "")
    expira_txt = f" — expira {expira[:10]}" if expira and premium else ""

    if not historico:
        conteudo = f"""
        <a class="back" href="/admin">← Voltar</a>
        <div class="card"><p>Nenhuma conversa para {telefone}.</p></div>"""
        return HTMLResponse(base_html(telefone, conteudo))

    msgs = ""
    for msg in historico:
        role  = msg["role"]
        texto = msg["content"].replace("\n", "<br>")
        if role == "user":
            msgs += f'<div class="msg usuario"><div class="label">Usuario</div><div class="balao">{texto}</div></div>'
        else:
            msgs += f'<div class="msg agente"><div class="label">{AGENT_NAME}</div><div class="balao">{texto}</div></div>'

    conteudo = f"""
    <a class="back" href="/admin">← Voltar</a>
    <div class="card" style="margin-bottom:12px;">
        <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;">
            <div>{status_badge}{expira_txt}</div>
            <div style="display:flex;gap:8px;">
                {"" if premium else f'<a href="/admin/assinaturas/ativar/{telefone}" class="btn btn-success">Ativar Premium</a>'}
                {f'<a href="/admin/assinaturas/desativar/{telefone}" onclick="return confirm(\'Desativar {telefone}?\')" class="btn btn-warning">Desativar</a>' if premium else ""}
            </div>
        </div>
    </div>
    <div class="card">
        <h2>Conversa com {telefone}</h2>
        <div class="total">{len(historico)} mensagens</div>
        <div class="chat">{msgs}</div>
    </div>"""

    return HTMLResponse(base_html(telefone, conteudo))


@app.get("/admin/apagar/{telefone}")
def apagar_historico(telefone: str, admin: str = Depends(verificar_admin)):
    r.delete(f"historico:{telefone}")
    return RedirectResponse(url="/admin")

# ============================================================
# PAINEL ADMIN — ASSINATURAS
# ============================================================

@app.get("/admin/assinaturas", response_class=HTMLResponse)
def painel_assinaturas(admin: str = Depends(verificar_admin), msg: str = ""):
    assinaturas = listar_assinaturas()
    aviso = f'<div class="success">{msg}</div>' if msg else ""

    rows = ""
    for assin in assinaturas:
        tel     = assin.get("telefone", "")
        status  = assin.get("status", "freemium")
        plano   = assin.get("plano", "freemium")
        expira  = assin.get("expira", "")[:10] if assin.get("expira") else "—"
        origem  = assin.get("origem", assin.get("status_mp", "manual"))

        if status == "ativo":
            badge = '<span class="badge badge-ativo">Ativo</span>'
        elif status == "inativo":
            badge = '<span class="badge badge-inativo">Inativo</span>'
        else:
            badge = '<span class="badge badge-freemium">Freemium</span>'

        btns = ""
        if status != "ativo":
            btns += f'<a href="/admin/assinaturas/ativar/{tel}" class="btn btn-success" style="margin-right:6px">Ativar</a>'
        if status == "ativo":
            btns += f'<a href="/admin/assinaturas/desativar/{tel}" onclick="return confirm(\'Desativar {tel}?\')" class="btn btn-warning" style="margin-right:6px">Desativar</a>'
        btns += f'<a href="/admin/assinaturas/apagar/{tel}" onclick="return confirm(\'Remover {tel}?\')" class="btn btn-danger">Remover</a>'

        rows += f"""
        <div class="aluno-row">
            <div>
                <div><strong>{tel}</strong> {badge} <span class="badge" style="background:#e0e7ff;color:#4f46e5">{plano}</span></div>
                <div class="aluno-info">Expira: {expira} | Origem: {origem}</div>
            </div>
            <div style="display:flex;gap:6px;flex-wrap:wrap">{btns}</div>
        </div>"""

    if not rows:
        rows = "<p style='color:#888;padding:20px 0'>Nenhuma assinatura ainda. Elas aparecem aqui apos o primeiro pagamento ou ativacao manual.</p>"

    conteudo = f"""
    {aviso}
    <div class="card">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
            <h2>Assinaturas ({len(assinaturas)})</h2>
            <a href="/admin/assinaturas/nova" class="btn btn-primary">+ Ativar manualmente</a>
        </div>
        {rows}
    </div>"""

    return HTMLResponse(base_html("Assinaturas", conteudo, "assinaturas"))


@app.get("/admin/assinaturas/nova", response_class=HTMLResponse)
def nova_assinatura_get(admin: str = Depends(verificar_admin)):
    conteudo = """
    <a class="back" href="/admin/assinaturas">← Voltar</a>
    <div class="card">
        <h2>Ativar assinatura manualmente</h2>
        <p style="font-size:13px;color:#888;margin-bottom:16px;">
            Use para ativar um usuario que pagou por fora, por cortesia ou para testes.
        </p>
        <form method="post" action="/admin/assinaturas/nova">
            <div style="margin-bottom:12px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Telefone (com DDI, ex: 5511999998888)</label>
                <input type="text" name="telefone" required placeholder="5511999998888">
            </div>
            <div style="margin-bottom:12px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Dias de acesso</label>
                <input type="number" name="dias" value="30" min="1" max="365" style="width:120px;padding:10px;border:1px solid #ddd;border-radius:8px;">
            </div>
            <button type="submit" class="btn btn-primary">Ativar acesso</button>
        </form>
    </div>"""
    return HTMLResponse(base_html("Nova Assinatura", conteudo, "assinaturas"))


@app.post("/admin/assinaturas/nova")
async def nova_assinatura_post(
    telefone: str = Form(...),
    dias: int = Form(30),
    admin: str = Depends(verificar_admin)
):
    dados = {
        "telefone": telefone,
        "status": "ativo",
        "plano": "premium",
        "origem": "manual",
        "data_inicio": datetime.now().isoformat(),
        "expira": (datetime.now() + timedelta(days=dias)).isoformat()
    }
    salvar_assinatura(telefone, dados)
    print(f"ASSINATURA MANUAL: {telefone} por {dias} dias")
    return RedirectResponse(url="/admin/assinaturas?msg=Assinatura+ativada+com+sucesso!", status_code=303)


@app.get("/admin/assinaturas/ativar/{telefone}")
def ativar_assinatura(telefone: str, admin: str = Depends(verificar_admin)):
    assinatura = obter_assinatura(telefone)
    assinatura["status"] = "ativo"
    assinatura["plano"]  = "premium"
    assinatura["expira"] = (datetime.now() + timedelta(days=35)).isoformat()
    assinatura["origem"] = assinatura.get("origem", "manual")
    salvar_assinatura(telefone, assinatura)
    return RedirectResponse(url="/admin/assinaturas?msg=Usuario+ativado!")


@app.get("/admin/assinaturas/desativar/{telefone}")
def desativar_assinatura(telefone: str, admin: str = Depends(verificar_admin)):
    assinatura = obter_assinatura(telefone)
    assinatura["status"] = "inativo"
    salvar_assinatura(telefone, assinatura)
    return RedirectResponse(url="/admin/assinaturas?msg=Usuario+desativado!")


@app.get("/admin/assinaturas/apagar/{telefone}")
def apagar_assinatura(telefone: str, admin: str = Depends(verificar_admin)):
    r.delete(f"{ASSINATURA_PREFIX}{telefone}")
    return RedirectResponse(url="/admin/assinaturas")

# ============================================================
# PAINEL ADMIN — CONSULTAS
# ============================================================

@app.get("/admin/consultas", response_class=HTMLResponse)
def painel_consultas(admin: str = Depends(verificar_admin)):
    consultas = listar_consultas()
    pendentes = [c for c in consultas if not c.get("atendido")]
    atendidas = [c for c in consultas if c.get("atendido")]

    def render_rows(lista, mostrar_btn=True):
        if not lista:
            return "<p style='color:#888;padding:12px 0'>Nenhuma.</p>"
        rows = ""
        for c in lista:
            tel  = c.get("telefone", "")
            nome = c.get("nome", "—")
            data = c.get("data", "—")
            btn  = f'<a href="/admin/consultas/atender/{tel}" class="btn btn-success">Marcar como atendido</a>' if mostrar_btn else '<span style="font-size:12px;color:#16a34a">✓ Atendido</span>'
            rows += f"""
            <div class="aluno-row">
                <div>
                    <div><strong>{tel}</strong></div>
                    <div class="aluno-info">Mensagem: {nome} | Data: {data}</div>
                </div>
                {btn}
            </div>"""
        return rows

    conteudo = f"""
    <div class="card">
        <h2>Interessados em Consulta — Pendentes ({len(pendentes)})</h2>
        {render_rows(pendentes, True)}
    </div>
    <div class="card">
        <h2>Ja Atendidos ({len(atendidas)})</h2>
        {render_rows(atendidas, False)}
    </div>"""

    return HTMLResponse(base_html("Consultas", conteudo, "consultas"))


@app.get("/admin/consultas/atender/{telefone}")
def atender_consulta(telefone: str, admin: str = Depends(verificar_admin)):
    marcar_consulta_atendida(telefone)
    return RedirectResponse(url="/admin/consultas")

# ============================================================
# PAINEL ADMIN — EDITAR PROMPT
# ============================================================

@app.get("/admin/prompt", response_class=HTMLResponse)
def editar_prompt_get(admin: str = Depends(verificar_admin), salvo: str = ""):
    prompt_atual = obter_prompt()
    aviso = '<div class="success">Prompt salvo com sucesso!</div>' if salvo == "1" else ""
    conteudo = f"""
    {aviso}
    <div class="card">
        <h2>Editar Prompt do Agente</h2>
        <p style="font-size:13px;color:#888;margin-bottom:16px;">
            Use <code>{{STATUS}}</code> para injetar FREEMIUM ou PREMIUM dinamicamente.<br>
            Use <code>[LINK_PAGAMENTO]</code> para injetar o link de assinatura.
        </p>
        <form method="post" action="/admin/prompt">
            <textarea name="prompt">{prompt_atual}</textarea>
            <br><br>
            <button type="submit" class="btn btn-primary">Salvar Prompt</button>
        </form>
    </div>"""
    return HTMLResponse(base_html("Editar Prompt", conteudo, "prompt"))


@app.post("/admin/prompt")
async def editar_prompt_post(
    prompt: str = Form(...),
    admin: str = Depends(verificar_admin)
):
    salvar_prompt(prompt.strip())
    return RedirectResponse(url="/admin/prompt?salvo=1", status_code=303)

# ============================================================
# PAINEL ADMIN — ARQUIVOS DE REFERENCIA
# ============================================================

@app.get("/admin/arquivos", response_class=HTMLResponse)
async def painel_arquivos(admin: str = Depends(verificar_admin), salvo: str = ""):
    arquivos = listar_arquivos()
    aviso = '<div class="success">Arquivo salvo com sucesso!</div>' if salvo == "1" else ""

    rows = ""
    for arq in arquivos:
        kb = round(arq["tamanho"] / 1024, 1)
        nome_arq = arq["nome"]
        rows += f"""
        <div class="aluno-row">
            <div>
                <div><strong>[{nome_arq}]</strong> &nbsp;<span style="font-size:12px;color:#888;">{kb} KB</span></div>
                <div class="aluno-info">Use [{nome_arq}] no prompt para referenciar este arquivo</div>
            </div>
            <a href="/admin/arquivos/apagar/{nome_arq}" onclick="return confirm('Apagar {nome_arq}?')">
                <span class="btn btn-danger">Apagar</span>
            </a>
        </div>"""

    if not rows:
        rows = "<p style='color:#888;padding:12px 0'>Nenhum arquivo ainda.</p>"

    conteudo = f"""
    {aviso}
    <div class="card">
        <h2>Arquivos de Referencia</h2>
        <p style="font-size:13px;color:#888;margin-bottom:16px;">
            Faca upload de PDFs ou arquivos de texto. Depois use <code>[nome]</code> no prompt.
        </p>
        {rows}
    </div>
    <div class="card">
        <h2>Novo Arquivo</h2>
        <form method="post" action="/admin/arquivos" enctype="multipart/form-data">
            <div style="margin-bottom:12px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Nome de referencia (sem espacos)</label>
                <input type="text" name="nome" required placeholder="metodologia" style="width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px;">
            </div>
            <div style="margin-bottom:16px;">
                <label style="font-size:13px;color:#555;display:block;margin-bottom:6px;">Arquivo (PDF ou TXT — max 500KB)</label>
                <input type="file" name="arquivo" accept=".pdf,.txt,.md" required style="font-size:14px;">
            </div>
            <button type="submit" class="btn btn-primary">Salvar Arquivo</button>
        </form>
    </div>"""

    return HTMLResponse(base_html("Arquivos", conteudo, "arquivos"))


@app.post("/admin/arquivos")
async def upload_arquivo(
    nome: str = Form(...),
    arquivo: UploadFile = File(...),
    admin: str = Depends(verificar_admin)
):
    conteudo_bytes = await arquivo.read()
    filename = arquivo.filename.lower()

    if filename.endswith(".pdf"):
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(conteudo_bytes)) as pdf:
                paginas = [p.extract_text() for p in pdf.pages[:30] if p.extract_text()]
            texto_final = "\n\n".join(paginas)
        except Exception as e:
            texto_final = f"[Erro ao ler PDF: {e}]"
    else:
        try:
            texto_final = conteudo_bytes.decode("utf-8")
        except Exception:
            texto_final = conteudo_bytes.decode("latin-1", errors="ignore")

    nome_seguro = re.sub(r"[^a-zA-Z0-9_\-]", "", nome).lower() or "arquivo"
    salvar_arquivo(nome_seguro, texto_final)
    return RedirectResponse(url="/admin/arquivos?salvo=1", status_code=303)


@app.get("/admin/arquivos/apagar/{nome}")
def apagar_arquivo_rota(nome: str, admin: str = Depends(verificar_admin)):
    apagar_arquivo(nome)
    return RedirectResponse(url="/admin/arquivos")

# ============================================================
# WEBHOOK Z-API (WhatsApp)
# ============================================================

@app.post("/webhook")
async def webhook(request: Request):
    try:
        dados = await request.json()

        if dados.get("type") != "ReceivedCallback":
            return {"status": "ignorado"}
        if dados.get("fromMe"):
            return {"status": "ignorado"}
        if dados.get("isGroup"):
            return {"status": "ignorado"}

        telefone = dados.get("phone", "")
        if not telefone:
            return {"status": "ignorado"}

        texto_midia = await processar_midia(dados)
        if texto_midia:
            resposta = await chamar_claude(telefone, texto_midia)
            await enviar_whatsapp(telefone, resposta)
            return {"status": "ok"}

        texto = dados.get("text", {}).get("message", "")
        if not texto:
            return {"status": "ignorado"}

        print(f"MSG de {telefone}: {texto[:80]}")
        resposta = await chamar_claude(telefone, texto)
        await enviar_whatsapp(telefone, resposta)
        return {"status": "ok"}

    except Exception as e:
        print(f"ERRO webhook: {e}")
        return {"status": "erro", "detalhe": str(e)}
